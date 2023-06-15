from django.shortcuts import render, redirect, HttpResponse

from manage_app.models import *


def depart_list(request):
    """部门列表"""
    queryset = Department.objects.all()
    return render(request, 'depart_list.html', {'data_list': queryset})


def add_depart(request):
    """添加部门"""
    if request.method == "GET":
        return render(request, 'add_depart.html')
    data = request.POST
    title = data.get("title")
    Department.objects.create(title=title)
    return redirect('/depart/list/')


def delete_depart(request):
    """删除部门"""
    nid = request.GET.get("id")
    Department.objects.filter(id=nid).delete()
    return redirect('/depart/list')


def edit_depart(request, nid):
    """编辑部门"""
    if request.method == "GET":
        obj = Department.objects.filter(id=nid).first()
        return render(request, 'edit_depart.html', {"obj": obj})
    title = request.POST.get("title")
    Department.objects.filter(id=nid).update(title=title)
    return redirect('/depart/list/')


def depart_multi(request):
    from openpyxl import load_workbook

    # 1.获取文件对象
    file_obj = request.FILES.get("tfile")

    # 2.将对象传递给 openpyxl，由openpyxl读取文件内容
    wb = load_workbook(file_obj)
    # 2.1 读取第1个 sheet 的内容
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exist = Department.objects.filter(title=text).exists()
        if not exist:
            Department.objects.create(title=text)

    return redirect('/depart/list/')